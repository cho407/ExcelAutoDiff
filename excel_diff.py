#!/usr/bin/env python3
from __future__ import annotations

import argparse
from concurrent.futures import ThreadPoolExecutor
import hashlib
import json
import math
import os
import re
import sys
import tempfile
import unicodedata
from copy import copy
from dataclasses import dataclass, field
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.formula.tokenizer import Tokenizer
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet
from zipfile import BadZipFile, ZipFile


SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm"}
CELL_COORD_RE = re.compile(r"^([A-Z]+)([0-9]+)$")
MODIFIED_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
ADDED_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")
REMOVED_FILL = PatternFill(fill_type="solid", fgColor="F8CBAD")
BASE_WHITE_FILL = PatternFill(fill_type="solid", fgColor="FFFFFF")
BLACK_FONT_COLOR = "000000"
CHANGED_TAB_COLOR = "FFD966"
ADDED_TAB_COLOR = "A9D08E"
DIFF_SUMMARY_TEXT_MAX_LEN = 600
EXCEL_MAX_ROWS = 1_048_576
INVISIBLE_TEXT_CHARS = {
    "\u200b",  # zero width space
    "\u200c",  # zero width non-joiner
    "\u200d",  # zero width joiner
    "\u2060",  # word joiner
    "\ufeff",  # bom / zero width no-break space
    "\u00ad",  # soft hyphen
}
NBSP_TEXT_CHARS = {
    "\u00a0",  # no-break space
    "\u202f",  # narrow no-break space
    "\u2007",  # figure space
}
SHEET_REF_QUOTED_RE = re.compile(r"'((?:''|[^'])+)'!")


@dataclass(frozen=True)
class CellSnapshot:
    compare_key: tuple[Any, ...]
    display_value: str


@dataclass
class CellChange:
    change_type: str
    address: str
    before: str | None = None
    after: str | None = None


@dataclass
class SheetDiff:
    sheet_name: str
    added_cells: int = 0
    removed_cells: int = 0
    modified_cells: int = 0
    details: list[CellChange] = field(default_factory=list)
    details_truncated: bool = False
    before_sheet_state: str | None = None
    after_sheet_state: str | None = None

    @property
    def has_changes(self) -> bool:
        return (
            self.added_cells > 0
            or self.removed_cells > 0
            or self.modified_cells > 0
            or self.before_sheet_state != self.after_sheet_state
        )


@dataclass
class FileDiff:
    relative_path: str
    left_path: str | None = None
    right_path: str | None = None
    status: str = "unchanged"  # unchanged|changed|added|removed|error
    added_sheets: list[str] = field(default_factory=list)
    removed_sheets: list[str] = field(default_factory=list)
    sheet_diffs: list[SheetDiff] = field(default_factory=list)
    error: str | None = None

    @property
    def changed_cells(self) -> int:
        return sum(sd.added_cells + sd.removed_cells + sd.modified_cells for sd in self.sheet_diffs)

    @property
    def has_changes(self) -> bool:
        if self.status in {"changed", "added", "removed", "error"}:
            return True
        if self.added_sheets or self.removed_sheets:
            return True
        return any(sd.has_changes for sd in self.sheet_diffs)


@dataclass
class DiffReport:
    left_target: str
    right_target: str
    mode: str  # file|directory
    scan_errors: list[str] = field(default_factory=list)
    file_diffs: list[FileDiff] = field(default_factory=list)

    @property
    def changed_files(self) -> int:
        return sum(1 for f in self.file_diffs if f.has_changes and f.status != "error")

    @property
    def error_files(self) -> int:
        return sum(1 for f in self.file_diffs if f.status == "error")

    @property
    def unchanged_files(self) -> int:
        return sum(1 for f in self.file_diffs if f.status == "unchanged")

    @property
    def modified_files(self) -> int:
        count = 0
        for fd in self.file_diffs:
            if any(
                sd.modified_cells > 0 or sd.before_sheet_state != sd.after_sheet_state
                for sd in fd.sheet_diffs
            ):
                count += 1
        return count


@dataclass
class HighlightSummary:
    modified_cells: int = 0
    added_cells: int = 0
    removed_cells: int = 0
    changed_sheets: list[str] = field(default_factory=list)
    added_sheets: list[str] = field(default_factory=list)
    removed_sheets: list[str] = field(default_factory=list)
    diff_rows: list[tuple[str, str, str, str, str]] = field(default_factory=list)


def is_supported_excel(path: Path) -> bool:
    return path.suffix.lower() in SUPPORTED_EXTENSIONS


def sha256_file(path: Path) -> str:
    hasher = hashlib.sha256()
    with path.open("rb") as f:
        while True:
            chunk = f.read(1024 * 1024)
            if not chunk:
                break
            hasher.update(chunk)
    return hasher.hexdigest()


def normalize_text_for_compare(value: str) -> str:
    text = unicodedata.normalize("NFKC", value)
    text = text.replace("\r\n", "\n").replace("\r", "\n")

    for ch in NBSP_TEXT_CHARS:
        text = text.replace(ch, " ")
    for ch in INVISIBLE_TEXT_CHARS:
        text = text.replace(ch, "")

    # 연속 공백/줄바꿈을 하나의 공백으로 통일해 보기상 동일한 값은 변경으로 보지 않는다.
    text = re.sub(r"\s+", " ", text).strip()
    return text


def can_unquote_sheet_name(name: str) -> bool:
    if not name:
        return False
    return re.search(r"[\s\[\]\*\?/\\:']", name) is None


def normalize_formula_range_token(value: str) -> str:
    token_text = value.strip()

    def replace_sheet_ref(match: re.Match[str]) -> str:
        raw = match.group(1)
        unescaped = raw.replace("''", "'")
        if can_unquote_sheet_name(unescaped):
            return f"{unescaped}!"
        return f"'{raw}'!"

    token_text = SHEET_REF_QUOTED_RE.sub(replace_sheet_ref, token_text)

    # 동적 배열 호환성으로 자동 삽입되는 @는 의미 차이 없는 경우가 많아 비교에서 제외한다.
    token_text = re.sub(r"^@+", "", token_text)
    return token_text.upper()


def normalize_formula_for_compare(value: str) -> str:
    text = normalize_text_for_compare(value)
    if not text.startswith("="):
        text = "=" + text

    try:
        tokens = Tokenizer(text).items
    except Exception:  # noqa: BLE001
        # 토크나이징 실패 시 공백/대소문자 정규화만 적용한 fallback
        return text[1:].upper()

    normalized_parts: list[str] = []

    for token in tokens:
        if token.type == "WHITE-SPACE":
            continue

        if token.type == "OPERAND" and token.subtype == "RANGE":
            normalized_parts.append(normalize_formula_range_token(token.value))
        elif token.type == "OPERAND" and token.subtype in {"LOGICAL", "ERROR"}:
            normalized_parts.append(token.value.upper())
        elif token.type == "FUNC" and token.subtype == "OPEN":
            normalized_parts.append(token.value[:-1].strip().upper() + "(")
        else:
            normalized_parts.append(token.value)

    return "".join(normalized_parts)


def normalize_value(value: Any) -> Any:
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, str):
        return normalize_text_for_compare(value)
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        # 1(int) vs 1.0(float) 같은 숫자를 동일 값으로 비교하기 위해 문자열로 통일
        return str(value)
    if isinstance(value, float):
        if math.isnan(value):
            return "NaN"
        if math.isinf(value):
            return "Infinity" if value > 0 else "-Infinity"
        return format(value, ".15g")
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")
    if value is None:
        return value
    return str(value)


def display_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, float):
        if math.isnan(value):
            return "NaN"
        if math.isinf(value):
            return "Infinity" if value > 0 else "-Infinity"
        return format(value, ".15g")
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")
    return str(value)


def make_cell_snapshot(cell: Any, include_format: bool) -> CellSnapshot | None:
    if isinstance(cell, MergedCell):
        return None
    if cell.value is None:
        return None

    if cell.data_type == "f" and isinstance(cell.value, str):
        normalized_value = normalize_formula_for_compare(cell.value)
        key: tuple[Any, ...] = ("formula", normalized_value)
        if include_format:
            key = key + (cell.number_format,)
        return CellSnapshot(compare_key=key, display_value=display_value(cell.value))

    normalized_value = normalize_value(cell.value)
    if isinstance(cell.value, str) and normalized_value == "":
        return None

    key: tuple[Any, ...] = (cell.data_type, normalized_value)
    if include_format:
        key = key + (cell.number_format,)
    return CellSnapshot(compare_key=key, display_value=display_value(cell.value))


def extract_cells(ws: Worksheet, include_format: bool) -> dict[str, CellSnapshot]:
    cells: dict[str, CellSnapshot] = {}
    internal_cells = getattr(ws, "_cells", None)
    if internal_cells:
        for row, col in sorted(internal_cells.keys()):
            cell = internal_cells[(row, col)]
            snapshot = make_cell_snapshot(cell, include_format=include_format)
            if snapshot is None:
                continue
            coord = f"{get_column_letter(col)}{row}"
            cells[coord] = snapshot
        return cells

    for row in ws.iter_rows():
        for cell in row:
            snapshot = make_cell_snapshot(cell, include_format=include_format)
            if snapshot is None:
                continue
            cells[cell.coordinate] = snapshot
    return cells


def normalize_validation_formula(value: Any) -> str:
    if value is None:
        return ""
    return normalize_text_for_compare(str(value))


def list_validation_signature(validation: Any) -> tuple[Any, ...]:
    return (
        normalize_text_for_compare(str(getattr(validation, "type", "") or "")),
        normalize_validation_formula(getattr(validation, "formula1", "")),
        normalize_validation_formula(getattr(validation, "formula2", "")),
        normalize_text_for_compare(str(getattr(validation, "operator", "") or "")),
        bool(getattr(validation, "allowBlank", False)),
        bool(getattr(validation, "showDropDown", False)),
    )


def get_list_validation_signatures_for_cell(
    ws: Worksheet,
    address: str,
    cache: dict[str, tuple[tuple[Any, ...], ...]],
) -> tuple[tuple[Any, ...], ...]:
    if address in cache:
        return cache[address]

    signatures: set[tuple[Any, ...]] = set()
    container = getattr(ws, "data_validations", None)
    validation_items = getattr(container, "dataValidation", []) if container else []

    for validation in validation_items:
        if str(getattr(validation, "type", "")).lower() != "list":
            continue
        try:
            if address in validation:
                signatures.add(list_validation_signature(validation))
        except Exception:  # noqa: BLE001
            continue

    result = tuple(sorted(signatures))
    cache[address] = result
    return result


def should_ignore_dropdown_selection_change(
    left_ws: Worksheet,
    right_ws: Worksheet,
    address: str,
    left_cache: dict[str, tuple[tuple[Any, ...], ...]],
    right_cache: dict[str, tuple[tuple[Any, ...], ...]],
) -> bool:
    left_signatures = get_list_validation_signatures_for_cell(left_ws, address, left_cache)
    if not left_signatures:
        return False
    right_signatures = get_list_validation_signatures_for_cell(right_ws, address, right_cache)
    if not right_signatures:
        return False
    return left_signatures == right_signatures


def cell_sort_key(coord: str) -> tuple[int, int]:
    match = CELL_COORD_RE.match(coord)
    if not match:
        return (sys.maxsize, sys.maxsize)
    col_text, row_text = match.groups()
    col_idx = 0
    for c in col_text:
        col_idx = col_idx * 26 + (ord(c) - ord("A") + 1)
    return (int(row_text), col_idx)


def load_workbook_safely(path: Path):
    try:
        return load_workbook(
            filename=str(path),
            data_only=False,
            read_only=False,
            keep_links=False,
            keep_vba=path.suffix.lower() == ".xlsm",
        ), None
    except PermissionError:
        return None, f"권한 문제 또는 잠금 상태로 파일에 접근할 수 없습니다: {path}"
    except BadZipFile:
        return None, f"엑셀 파일 구조를 읽을 수 없습니다(손상/암호화 가능성): {path}"
    except Exception as exc:  # noqa: BLE001
        return None, f"파일 로드 실패: {path} ({exc})"


def compare_sheet(
    left_ws: Worksheet,
    right_ws: Worksheet,
    include_format: bool,
    max_cell_details: int,
) -> SheetDiff:
    left_cells = extract_cells(left_ws, include_format=include_format)
    right_cells = extract_cells(right_ws, include_format=include_format)

    left_keys = set(left_cells.keys())
    right_keys = set(right_cells.keys())

    raw_added_keys = sorted(right_keys - left_keys, key=cell_sort_key)
    raw_removed_keys = sorted(left_keys - right_keys, key=cell_sort_key)
    common_keys = sorted(left_keys & right_keys, key=cell_sort_key)

    left_dv_cache: dict[str, tuple[tuple[Any, ...], ...]] = {}
    right_dv_cache: dict[str, tuple[tuple[Any, ...], ...]] = {}

    added_keys = [
        addr
        for addr in raw_added_keys
        if not should_ignore_dropdown_selection_change(
            left_ws, right_ws, addr, left_dv_cache, right_dv_cache
        )
    ]
    removed_keys = [
        addr
        for addr in raw_removed_keys
        if not should_ignore_dropdown_selection_change(
            left_ws, right_ws, addr, left_dv_cache, right_dv_cache
        )
    ]

    modified_keys = [
        addr
        for addr in common_keys
        if left_cells[addr].compare_key != right_cells[addr].compare_key
        and not should_ignore_dropdown_selection_change(
            left_ws, right_ws, addr, left_dv_cache, right_dv_cache
        )
    ]

    diff = SheetDiff(
        sheet_name=left_ws.title,
        added_cells=len(added_keys),
        removed_cells=len(removed_keys),
        modified_cells=len(modified_keys),
        before_sheet_state=left_ws.sheet_state,
        after_sheet_state=right_ws.sheet_state,
    )

    total_details = 0
    for addr in added_keys:
        total_details += 1
        if len(diff.details) < max_cell_details:
            diff.details.append(
                CellChange(change_type="added", address=addr, before=None, after=right_cells[addr].display_value)
            )

    for addr in removed_keys:
        total_details += 1
        if len(diff.details) < max_cell_details:
            diff.details.append(
                CellChange(change_type="removed", address=addr, before=left_cells[addr].display_value, after=None)
            )

    for addr in modified_keys:
        total_details += 1
        if len(diff.details) < max_cell_details:
            diff.details.append(
                CellChange(
                    change_type="modified",
                    address=addr,
                    before=left_cells[addr].display_value,
                    after=right_cells[addr].display_value,
                )
            )

    diff.details_truncated = total_details > len(diff.details)
    return diff


def compare_workbooks(
    left_file: Path,
    right_file: Path,
    include_format: bool,
    max_cell_details: int,
    relative_path: str,
) -> FileDiff:
    file_diff = FileDiff(relative_path=relative_path, left_path=str(left_file), right_path=str(right_file))

    left_wb, left_error = load_workbook_safely(left_file)
    if left_error:
        file_diff.status = "error"
        file_diff.error = left_error
        return file_diff

    right_wb, right_error = load_workbook_safely(right_file)
    if right_error:
        file_diff.status = "error"
        file_diff.error = right_error
        if left_wb:
            left_wb.close()
        return file_diff

    try:
        left_sheets = set(left_wb.sheetnames)
        right_sheets = set(right_wb.sheetnames)
        file_diff.added_sheets = sorted(right_sheets - left_sheets)
        file_diff.removed_sheets = sorted(left_sheets - right_sheets)

        common_sheets = sorted(left_sheets & right_sheets)
        for sheet_name in common_sheets:
            sheet_diff = compare_sheet(
                left_wb[sheet_name],
                right_wb[sheet_name],
                include_format=include_format,
                max_cell_details=max_cell_details,
            )
            if sheet_diff.has_changes:
                file_diff.sheet_diffs.append(sheet_diff)

        if file_diff.added_sheets or file_diff.removed_sheets or file_diff.sheet_diffs:
            file_diff.status = "changed"
        else:
            file_diff.status = "unchanged"
    finally:
        left_wb.close()
        right_wb.close()

    return file_diff


def scan_excel_files(root: Path) -> tuple[dict[str, Path], list[str]]:
    found: dict[str, Path] = {}
    scan_errors: list[str] = []

    def on_error(err: OSError) -> None:
        filename = err.filename or str(root)
        scan_errors.append(f"탐색 실패: {filename} ({err.strerror})")

    for dir_path, _dir_names, filenames in os.walk(root, onerror=on_error):
        dir_obj = Path(dir_path)
        for file_name in filenames:
            if file_name.startswith("~$"):
                continue
            candidate = dir_obj / file_name
            if not is_supported_excel(candidate):
                continue
            rel = str(candidate.relative_to(root))
            found[rel] = candidate
    return found, scan_errors


def default_worker_count() -> int:
    cpu_count = os.cpu_count() or 4
    return max(1, min(8, cpu_count))


def compare_workbooks_task(
    left_file: Path,
    right_file: Path,
    include_format: bool,
    max_cell_details: int,
    relative_path: str,
) -> FileDiff:
    try:
        return compare_workbooks(
            left_file=left_file,
            right_file=right_file,
            include_format=include_format,
            max_cell_details=max_cell_details,
            relative_path=relative_path,
        )
    except Exception as exc:  # noqa: BLE001
        return FileDiff(
            relative_path=relative_path,
            left_path=str(left_file),
            right_path=str(right_file),
            status="error",
            error=f"비교 실패: {relative_path} ({exc})",
        )


def compare_file_pair_task(
    left_file: Path,
    right_file: Path,
    include_format: bool,
    max_cell_details: int,
    relative_path: str,
) -> tuple[FileDiff, list[str]]:
    scan_errors: list[str] = []

    try:
        same_size = left_file.stat().st_size == right_file.stat().st_size
        if same_size and sha256_file(left_file) == sha256_file(right_file):
            return (
                FileDiff(
                    relative_path=relative_path,
                    left_path=str(left_file),
                    right_path=str(right_file),
                    status="unchanged",
                ),
                scan_errors,
            )
    except OSError as exc:
        scan_errors.append(f"해시 계산 실패: {relative_path} ({exc})")
    except Exception as exc:  # noqa: BLE001
        return (
            FileDiff(
                relative_path=relative_path,
                left_path=str(left_file),
                right_path=str(right_file),
                status="error",
                error=f"비교 실패: {relative_path} ({exc})",
            ),
            scan_errors,
        )

    return (
        compare_workbooks_task(
            left_file=left_file,
            right_file=right_file,
            include_format=include_format,
            max_cell_details=max_cell_details,
            relative_path=relative_path,
        ),
        scan_errors,
    )


def compare_targets(
    left_target: Path,
    right_target: Path,
    include_format: bool,
    max_cell_details: int,
    workers: int = 1,
) -> DiffReport:
    left_abs = left_target.resolve()
    right_abs = right_target.resolve()

    if left_target.is_file() and right_target.is_file():
        if not is_supported_excel(left_target) or not is_supported_excel(right_target):
            raise ValueError("파일 비교 모드에서는 양쪽 모두 xlsx/xlsm 파일이어야 합니다.")

        report = DiffReport(
            left_target=str(left_abs),
            right_target=str(right_abs),
            mode="file",
        )
        file_diff = compare_workbooks(
            left_file=left_target,
            right_file=right_target,
            include_format=include_format,
            max_cell_details=max_cell_details,
            relative_path=f"{left_target.name} <-> {right_target.name}",
        )
        report.file_diffs.append(file_diff)
        return report

    if left_target.is_dir() and right_target.is_dir():
        report = DiffReport(
            left_target=str(left_abs),
            right_target=str(right_abs),
            mode="directory",
        )
        left_files, left_errors = scan_excel_files(left_target)
        right_files, right_errors = scan_excel_files(right_target)
        report.scan_errors.extend(left_errors)
        report.scan_errors.extend(right_errors)

        left_keys = set(left_files.keys())
        right_keys = set(right_files.keys())

        only_left = sorted(left_keys - right_keys)
        only_right = sorted(right_keys - left_keys)
        common = sorted(left_keys & right_keys)

        for rel in only_left:
            report.file_diffs.append(
                FileDiff(
                    relative_path=rel,
                    left_path=str(left_files[rel]),
                    right_path=None,
                    status="removed",
                )
            )

        for rel in only_right:
            report.file_diffs.append(
                FileDiff(
                    relative_path=rel,
                    left_path=None,
                    right_path=str(right_files[rel]),
                    status="added",
                )
            )

        compare_jobs: list[tuple[Path, Path, str]] = [
            (left_files[rel], right_files[rel], rel) for rel in common
        ]

        compare_results: list[tuple[FileDiff, list[str]]] = []
        if compare_jobs:
            def run_compare_job(job: tuple[Path, Path, str]) -> tuple[FileDiff, list[str]]:
                left_file, right_file, rel = job
                return compare_file_pair_task(
                    left_file=left_file,
                    right_file=right_file,
                    include_format=include_format,
                    max_cell_details=max_cell_details,
                    relative_path=rel,
                )

            if workers <= 1:
                compare_results = [run_compare_job(job) for job in compare_jobs]
            else:
                with ThreadPoolExecutor(max_workers=workers) as executor:
                    compare_results = list(executor.map(run_compare_job, compare_jobs))

        for file_diff, job_scan_errors in compare_results:
            report.scan_errors.extend(job_scan_errors)
            report.file_diffs.append(file_diff)

        report.file_diffs.sort(key=lambda x: x.relative_path)
        return report

    raise ValueError("입력 타입이 다릅니다. 파일 vs 파일 또는 폴더 vs 폴더 형태로 비교해야 합니다.")


def patch_workbook_xml_bytes(data: bytes) -> bytes:
    data = re.sub(rb"<workbookProtection[^>]*\/>", b"", data)
    data = re.sub(rb"<workbookProtection.*?</workbookProtection>", b"", data, flags=re.DOTALL)
    data = data.replace(b' state="veryHidden"', b"")
    data = data.replace(b' state="hidden"', b"")
    data = data.replace(b" state='veryHidden'", b"")
    data = data.replace(b" state='hidden'", b"")
    return data


def patch_worksheet_xml_bytes(data: bytes) -> bytes:
    data = re.sub(rb"<sheetProtection[^>]*\/>", b"", data)
    data = re.sub(rb"<sheetProtection.*?</sheetProtection>", b"", data, flags=re.DOTALL)
    data = data.replace(b' hidden="1"', b"")
    data = data.replace(b' hidden="true"', b"")
    data = data.replace(b" hidden='1'", b"")
    data = data.replace(b" hidden='true'", b"")
    return data


def create_unprotected_copy(input_file: Path, output_file: Path) -> None:
    with ZipFile(input_file, "r") as zin:
        with ZipFile(output_file, "w") as zout:
            for info in zin.infolist():
                name = info.filename
                if name.endswith(".DS_Store") or "/.DS_Store" in name:
                    continue

                data = zin.read(name)
                if name == "xl/workbook.xml":
                    data = patch_workbook_xml_bytes(data)
                elif name.startswith("xl/worksheets/") and name.endswith(".xml"):
                    data = patch_worksheet_xml_bytes(data)
                elif name == "xl/vbaProject.bin":
                    data = data.replace(b"DPB=", b"DPx=")

                zout.writestr(info, data)


def apply_fill_safe(ws: Worksheet, address: str, fill: PatternFill) -> bool:
    target = ws[address]
    if isinstance(target, MergedCell):
        return False
    target.fill = fill
    current_font = target.font
    if current_font is None:
        target.font = Font(color=BLACK_FONT_COLOR)
    else:
        cloned_font = copy(current_font)
        cloned_font.color = BLACK_FONT_COLOR
        target.font = cloned_font
    return True


def normalize_cell_visual_to_white_black(ws: Worksheet) -> None:
    # 사용 중인 셀들의 기존 강조 색/글자색을 제거해 가독성을 통일한다.
    internal_cells = getattr(ws, "_cells", {})
    for cell in internal_cells.values():
        if isinstance(cell, MergedCell):
            continue
        cell.fill = BASE_WHITE_FILL
        current_font = cell.font
        if current_font is None:
            cell.font = Font(color=BLACK_FONT_COLOR)
        else:
            cloned_font = copy(current_font)
            cloned_font.color = BLACK_FONT_COLOR
            cell.font = cloned_font

    ws.sheet_properties.tabColor = None

    # 기존 조건부서식 색상 영향을 제거해 변경 색상만 보이도록 정리한다.
    conditional_formatting = getattr(ws, "conditional_formatting", None)
    if conditional_formatting is not None:
        cf_rules = getattr(conditional_formatting, "_cf_rules", None)
        if cf_rules is not None:
            cf_rules.clear()


def normalize_workbook_visual_to_white_black(workbook: Any) -> None:
    for ws in workbook.worksheets:
        normalize_cell_visual_to_white_black(ws)


def compact_summary_text(value: str | None) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    text = text.replace("\n", " ⏎ ")
    if len(text) > DIFF_SUMMARY_TEXT_MAX_LEN:
        return text[:DIFF_SUMMARY_TEXT_MAX_LEN] + "..."
    return text


def excel_safe_text(value: str | None) -> str:
    text = compact_summary_text(value)
    if not text:
        return text

    # 요약 시트에서 수식/동적 링크로 재평가되지 않도록 문자열로 강제 저장한다.
    if text[0] in {"=", "+", "-", "@"} and not text.startswith("'"):
        return "'" + text
    return text


def add_diff_row(
    summary: HighlightSummary,
    change_type: str,
    sheet_name: str,
    cell_address: str,
    before: str | None,
    after: str | None,
) -> None:
    summary.diff_rows.append(
        (
            change_type,
            sheet_name,
            cell_address.upper(),
            before or "",
            after or "",
        )
    )


def set_summary_sheet_columns(ws: Worksheet) -> None:
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 42
    ws.column_dimensions["E"].width = 42


def write_diff_table_header(ws: Worksheet, title_row: int, header_row: int, title: str) -> None:
    ws[f"A{title_row}"] = title
    ws[f"A{header_row}"] = "구분"
    ws[f"B{header_row}"] = "시트"
    ws[f"C{header_row}"] = "셀"
    ws[f"D{header_row}"] = "이전 값"
    ws[f"E{header_row}"] = "변경 값"
    ws.freeze_panes = f"A{header_row + 1}"


def write_diff_rows_chunk(
    ws: Worksheet,
    rows: list[tuple[str, str, str, str, str]],
    start_row: int,
) -> int:
    if start_row > EXCEL_MAX_ROWS or not rows:
        return 0

    available = EXCEL_MAX_ROWS - start_row + 1
    write_count = min(len(rows), available)
    change_label = {"modified": "수정", "added": "추가", "removed": "삭제"}

    row_no = start_row
    for change_type, sheet_name, cell_addr, before_val, after_val in rows[:write_count]:
        ws[f"A{row_no}"] = change_label.get(change_type, change_type)
        ws[f"B{row_no}"] = sheet_name
        ws[f"C{row_no}"] = cell_addr
        ws[f"D{row_no}"] = excel_safe_text(before_val)
        ws[f"E{row_no}"] = excel_safe_text(after_val)
        row_no += 1
    return write_count


def apply_auto_filter_with_refresh(ws: Worksheet, ref: str) -> None:
    # 일부 엑셀 버전에서 필터 UI 초기화 타이밍이 늦는 현상을 줄이기 위해
    # 범위를 한 번 비웠다가 다시 지정해 메타데이터를 안정화한다.
    ws.auto_filter.ref = None
    ws.auto_filter.ref = ref


def collect_existing_table_names(workbook: Any) -> set[str]:
    names: set[str] = set()
    for ws in workbook.worksheets:
        tables = getattr(ws, "tables", None)
        if tables:
            try:
                names.update(str(name) for name in tables.keys())
            except Exception:  # noqa: BLE001
                pass
        legacy_tables = getattr(ws, "_tables", None)
        if legacy_tables:
            try:
                for table in legacy_tables:
                    table_name = getattr(table, "name", None)
                    if table_name:
                        names.add(str(table_name))
            except Exception:  # noqa: BLE001
                pass
    return names


def add_summary_table(ws: Worksheet, ref: str, preferred_name: str, reserved_names: set[str]) -> str:
    table_name = preferred_name
    suffix = 2
    while table_name in reserved_names:
        table_name = f"{preferred_name}_{suffix}"
        suffix += 1

    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight1",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    ws.add_table(table)
    reserved_names.add(table_name)
    return table_name


def enforce_single_selected_sheet(workbook: Any, selected_sheet_name: str) -> None:
    selected_index = 0
    for idx, ws in enumerate(workbook.worksheets):
        is_selected = ws.title == selected_sheet_name
        ws.sheet_view.tabSelected = bool(is_selected)
        if is_selected:
            selected_index = idx

    try:
        workbook.active = selected_index
    except Exception:  # noqa: BLE001
        pass

    if getattr(workbook, "views", None):
        for view in workbook.views:
            try:
                view.activeTab = selected_index
            except Exception:  # noqa: BLE001
                continue


def make_all_visible_and_unprotected(workbook: Any) -> None:
    if getattr(workbook, "security", None):
        workbook.security.lockStructure = False
        workbook.security.lockWindows = False
        workbook.security.lockRevision = False

    for ws in workbook.worksheets:
        ws.sheet_state = "visible"
        ws.protection.sheet = False
        ws.protection.objects = False
        ws.protection.scenarios = False
        for row_dim in ws.row_dimensions.values():
            row_dim.hidden = False
        for col_dim in ws.column_dimensions.values():
            col_dim.hidden = False


def write_highlight_summary_sheet(
    workbook: Any,
    summary: HighlightSummary,
    left_path: Path,
    right_path: Path,
) -> None:
    summary_name = "_DIFF_SUMMARY"
    existing_summary_sheets = [
        name
        for name in workbook.sheetnames
        if name == summary_name or name.startswith(f"{summary_name}_")
    ]
    for name in existing_summary_sheets:
        del workbook[name]

    reserved_table_names = collect_existing_table_names(workbook)

    ws = workbook.create_sheet(title=summary_name, index=0)
    ws["A1"] = "변경 하이라이트 안내"
    ws["A2"] = "비교 기준 파일"
    ws["B2"] = str(left_path.resolve())
    ws["A3"] = "색상 적용 파일(2번째 파일 기반)"
    ws["B3"] = str(right_path.resolve())

    ws["A5"] = "수정 셀"
    ws["A6"] = "추가 셀(2번째 파일에만 존재)"
    ws["A7"] = "삭제 셀(1번째 파일에만 존재)"
    ws["C5"] = summary.modified_cells
    ws["C6"] = summary.added_cells
    ws["C7"] = summary.removed_cells

    ws["A9"] = "시트 추가(2번째 파일)"
    ws["B9"] = ", ".join(summary.added_sheets) if summary.added_sheets else "-"
    ws["A10"] = "시트 삭제(1번째 파일)"
    ws["B10"] = ", ".join(summary.removed_sheets) if summary.removed_sheets else "-"
    ws["A11"] = "변경 감지 시트(탭 색상 표시)"
    ws["B11"] = ", ".join(summary.changed_sheets) if summary.changed_sheets else "-"
    ws["A13"] = "탭 색상 안내"
    ws["A14"] = "노란색 탭"
    ws["B14"] = "공통 시트에서 변경 감지"
    ws["A15"] = "초록색 탭"
    ws["B15"] = "2번째 파일에만 있는 신규 시트"
    ws["A16"] = "필터/정렬 안내"
    ws["B16"] = "헤더 행 필터에서 구분/시트/셀/이전 값/변경 값 기준 필터 가능"

    detail_title_row = 18
    detail_header_row = 19
    detail_data_row = 20
    write_diff_table_header(
        ws=ws,
        title_row=detail_title_row,
        header_row=detail_header_row,
        title="변경 상세 (시트/셀 기준)",
    )
    set_summary_sheet_columns(ws)

    written_count = write_diff_rows_chunk(
        ws=ws,
        rows=summary.diff_rows,
        start_row=detail_data_row,
    )
    if written_count > 0:
        add_summary_table(
            ws=ws,
            ref=f"A{detail_header_row}:E{detail_data_row + written_count - 1}",
            preferred_name="DiffSummaryTable",
            reserved_names=reserved_table_names,
        )
    else:
        apply_auto_filter_with_refresh(
            ws=ws,
            ref=f"A{detail_header_row}:E{detail_header_row}",
        )

    remaining_rows = summary.diff_rows[written_count:]

    page = 2
    while remaining_rows:
        continuation_name = f"{summary_name}_{page}"
        continuation_ws = workbook.create_sheet(title=continuation_name)
        continuation_ws["A1"] = "변경 상세 (계속)"
        continuation_ws["B1"] = f"{page} 페이지"
        write_diff_table_header(
            ws=continuation_ws,
            title_row=1,
            header_row=3,
            title="변경 상세 (시트/셀 기준)",
        )
        set_summary_sheet_columns(continuation_ws)
        chunk_written = write_diff_rows_chunk(
            ws=continuation_ws,
            rows=remaining_rows,
            start_row=4,
        )
        if chunk_written > 0:
            add_summary_table(
                ws=continuation_ws,
                ref=f"A3:E{3 + chunk_written}",
                preferred_name=f"DiffSummaryTable_{page}",
                reserved_names=reserved_table_names,
            )
        else:
            apply_auto_filter_with_refresh(ws=continuation_ws, ref="A3:E3")
        remaining_rows = remaining_rows[chunk_written:]
        page += 1

    if page > 2:
        ws["A17"] = "추가 상세 시트"
        ws["B17"] = f"{summary_name}_2 ~ {summary_name}_{page - 1}"

    enforce_single_selected_sheet(workbook=workbook, selected_sheet_name=summary_name)


def generate_highlight_workbook(
    left_file: Path,
    right_file: Path,
    output_file: Path,
    include_format: bool,
    unprotect_second: bool,
) -> HighlightSummary:
    summary = HighlightSummary()

    with tempfile.TemporaryDirectory(prefix="excel_diff_") as tmpdir:
        prepared_right = right_file
        if unprotect_second:
            prepared_right = Path(tmpdir) / f"{right_file.stem}_unprotected{right_file.suffix}"
            create_unprotected_copy(right_file, prepared_right)

        left_wb, left_error = load_workbook_safely(left_file)
        if left_error:
            raise RuntimeError(left_error)

        right_wb, right_error = load_workbook_safely(prepared_right)
        if right_error:
            if left_wb:
                left_wb.close()
            raise RuntimeError(right_error)

        try:
            make_all_visible_and_unprotected(right_wb)
            normalize_workbook_visual_to_white_black(right_wb)

            left_sheets = set(left_wb.sheetnames)
            right_sheets = set(right_wb.sheetnames)

            summary.added_sheets = sorted(right_sheets - left_sheets)
            summary.removed_sheets = sorted(left_sheets - right_sheets)
            changed_sheet_names: set[str] = set()

            for sheet_name in summary.added_sheets:
                right_ws = right_wb[sheet_name]
                right_ws.sheet_properties.tabColor = ADDED_TAB_COLOR
                changed_sheet_names.add(sheet_name)
                right_cells = extract_cells(right_ws, include_format=False)
                for addr in sorted(right_cells.keys(), key=cell_sort_key):
                    if apply_fill_safe(right_ws, addr, ADDED_FILL):
                        summary.added_cells += 1
                    add_diff_row(
                        summary=summary,
                        change_type="added",
                        sheet_name=sheet_name,
                        cell_address=addr,
                        before="",
                        after=right_cells[addr].display_value,
                    )

            for sheet_name in sorted(right_sheets & left_sheets):
                left_ws = left_wb[sheet_name]
                right_ws = right_wb[sheet_name]

                left_cells = extract_cells(left_ws, include_format=include_format)
                right_cells = extract_cells(right_ws, include_format=include_format)

                left_keys = set(left_cells.keys())
                right_keys = set(right_cells.keys())

                left_dv_cache: dict[str, tuple[tuple[Any, ...], ...]] = {}
                right_dv_cache: dict[str, tuple[tuple[Any, ...], ...]] = {}

                raw_added_keys = right_keys - left_keys
                raw_removed_keys = left_keys - right_keys
                raw_modified_keys = {
                    addr
                    for addr in (left_keys & right_keys)
                    if left_cells[addr].compare_key != right_cells[addr].compare_key
                }
                added_keys = {
                    addr
                    for addr in raw_added_keys
                    if not should_ignore_dropdown_selection_change(
                        left_ws, right_ws, addr, left_dv_cache, right_dv_cache
                    )
                }
                removed_keys = {
                    addr
                    for addr in raw_removed_keys
                    if not should_ignore_dropdown_selection_change(
                        left_ws, right_ws, addr, left_dv_cache, right_dv_cache
                    )
                }
                modified_keys = {
                    addr
                    for addr in raw_modified_keys
                    if not should_ignore_dropdown_selection_change(
                        left_ws, right_ws, addr, left_dv_cache, right_dv_cache
                    )
                }
                sheet_has_changes = bool(added_keys or removed_keys or modified_keys)

                for addr in sorted(added_keys, key=cell_sort_key):
                    if apply_fill_safe(right_ws, addr, ADDED_FILL):
                        summary.added_cells += 1
                    add_diff_row(
                        summary=summary,
                        change_type="added",
                        sheet_name=sheet_name,
                        cell_address=addr,
                        before="",
                        after=right_cells[addr].display_value,
                    )

                for addr in sorted(removed_keys, key=cell_sort_key):
                    if apply_fill_safe(right_ws, addr, REMOVED_FILL):
                        summary.removed_cells += 1
                    add_diff_row(
                        summary=summary,
                        change_type="removed",
                        sheet_name=sheet_name,
                        cell_address=addr,
                        before=left_cells[addr].display_value,
                        after="",
                    )

                for addr in sorted(modified_keys, key=cell_sort_key):
                    if apply_fill_safe(right_ws, addr, MODIFIED_FILL):
                        summary.modified_cells += 1
                    add_diff_row(
                        summary=summary,
                        change_type="modified",
                        sheet_name=sheet_name,
                        cell_address=addr,
                        before=left_cells[addr].display_value,
                        after=right_cells[addr].display_value,
                    )

                if sheet_has_changes:
                    right_ws.sheet_properties.tabColor = CHANGED_TAB_COLOR
                    changed_sheet_names.add(sheet_name)

            summary.changed_sheets = sorted(changed_sheet_names)

            write_highlight_summary_sheet(
                workbook=right_wb,
                summary=summary,
                left_path=left_file,
                right_path=right_file,
            )
            output_file.parent.mkdir(parents=True, exist_ok=True)
            right_wb.save(output_file)
        finally:
            left_wb.close()
            right_wb.close()

    return summary


def build_modified_views(report: DiffReport) -> list[dict[str, Any]]:
    modified_views: list[dict[str, Any]] = []

    for fd in report.file_diffs:
        if fd.status != "changed":
            continue

        sheet_views: list[dict[str, Any]] = []
        for sd in fd.sheet_diffs:
            state_changed = sd.before_sheet_state != sd.after_sheet_state
            if sd.modified_cells == 0 and not state_changed:
                continue

            modified_details = [d for d in sd.details if d.change_type == "modified"]
            details_truncated = sd.details_truncated or sd.modified_cells > len(modified_details)

            sheet_views.append(
                {
                    "sheet_name": sd.sheet_name,
                    "modified_cells": sd.modified_cells,
                    "before_sheet_state": sd.before_sheet_state,
                    "after_sheet_state": sd.after_sheet_state,
                    "details_truncated": details_truncated,
                    "details": modified_details,
                }
            )

        if not sheet_views:
            continue

        modified_views.append(
            {
                "relative_path": fd.relative_path,
                "left_path": fd.left_path,
                "right_path": fd.right_path,
                "sheet_views": sheet_views,
            }
        )

    return modified_views


def report_to_dict(report: DiffReport) -> dict[str, Any]:
    modified_views = build_modified_views(report)

    payload: dict[str, Any] = {
        "left_target": report.left_target,
        "right_target": report.right_target,
        "mode": report.mode,
        "modified_only_report": True,
        "summary": {
            "total_compared_files": len(report.file_diffs),
            "modified_files": len(modified_views),
            "error_files": report.error_files,
            "scan_errors": len(report.scan_errors),
        },
        "scan_errors": report.scan_errors,
        "files": [],
    }

    for view in modified_views:
        total_modified_cells = sum(sv["modified_cells"] for sv in view["sheet_views"])
        file_entry = {
            "relative_path": view["relative_path"],
            "left_path": view["left_path"],
            "right_path": view["right_path"],
            "modified_cells": total_modified_cells,
            "sheet_diffs": [],
        }
        for sheet_view in view["sheet_views"]:
            file_entry["sheet_diffs"].append(
                {
                    "sheet_name": sheet_view["sheet_name"],
                    "modified_cells": sheet_view["modified_cells"],
                    "before_sheet_state": sheet_view["before_sheet_state"],
                    "after_sheet_state": sheet_view["after_sheet_state"],
                    "details_truncated": sheet_view["details_truncated"],
                    "details": [
                        {
                            "change_type": d.change_type,
                            "address": d.address,
                            "before": d.before,
                            "after": d.after,
                        }
                        for d in sheet_view["details"]
                    ],
                }
            )
        payload["files"].append(file_entry)
    return payload


def render_markdown(report: DiffReport) -> str:
    modified_views = build_modified_views(report)

    lines: list[str] = []
    lines.append("# Excel 변경점 리포트")
    lines.append("")
    lines.append(f"- 비교 대상 A: `{report.left_target}`")
    lines.append(f"- 비교 대상 B: `{report.right_target}`")
    lines.append(f"- 비교 모드: `{report.mode}`")
    lines.append("")
    lines.append("## 요약")
    lines.append("")
    lines.append(f"- 전체 비교 파일: {len(report.file_diffs)}")
    lines.append(f"- 수정 파일: {len(modified_views)}")
    lines.append(f"- 오류 파일: {report.error_files}")
    lines.append(f"- 스캔 오류: {len(report.scan_errors)}")
    lines.append("")

    if report.scan_errors:
        lines.append("## 스캔 오류")
        lines.append("")
        for err in report.scan_errors:
            lines.append(f"- {err}")
        lines.append("")

    lines.append("## 파일별 상세")
    lines.append("")

    if not modified_views:
        lines.append("- 수정된 항목이 없습니다.")
        lines.append("")
        return "\n".join(lines)

    for view in modified_views:
        lines.append(f"### {view['relative_path']}")
        lines.append("")
        if view["left_path"]:
            lines.append(f"- A 경로: `{view['left_path']}`")
        if view["right_path"]:
            lines.append(f"- B 경로: `{view['right_path']}`")

        total_modified_cells = sum(sv["modified_cells"] for sv in view["sheet_views"])
        lines.append(f"- 수정 셀 수(합계): {total_modified_cells}")

        for sheet_view in view["sheet_views"]:
            lines.append(f"- 시트 `{sheet_view['sheet_name']}`: 수정 {sheet_view['modified_cells']}")
            if sheet_view["before_sheet_state"] != sheet_view["after_sheet_state"]:
                lines.append(
                    f"  - 시트 상태 변경: `{sheet_view['before_sheet_state']}` -> `{sheet_view['after_sheet_state']}`"
                )
            for detail in sheet_view["details"]:
                lines.append(
                    f"  - [수정] `{detail.address}`: `{detail.before}` -> `{detail.after}`"
                )
            if sheet_view["details_truncated"]:
                lines.append("  - (수정 상세 목록이 너무 많아 일부만 표시됨)")
        lines.append("")

    return "\n".join(lines)


def print_summary(report: DiffReport) -> None:
    modified_views = build_modified_views(report)

    print("=== Excel Diff Summary ===")
    print(f"Mode         : {report.mode}")
    print(f"Left target  : {report.left_target}")
    print(f"Right target : {report.right_target}")
    print(f"Total files  : {len(report.file_diffs)}")
    print(f"Modified only: {len(modified_views)}")
    print(f"Errors       : {report.error_files}")
    print(f"Scan errors  : {len(report.scan_errors)}")


def parse_args(argv: Iterable[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="xlsx/xlsm 파일 또는 폴더를 비교해 변경 셀/시트 정보를 리포트합니다."
    )
    parser.add_argument("left", help="기준 파일/폴더 경로")
    parser.add_argument("right", help="비교 대상 파일/폴더 경로")
    parser.add_argument(
        "--output",
        "-o",
        help="리포트 저장 경로(.md 또는 .json). 미지정 시 콘솔 요약만 출력",
    )
    parser.add_argument(
        "--format",
        choices=["md", "json"],
        default="md",
        help="리포트 포맷(기본: md)",
    )
    parser.add_argument(
        "--include-format",
        action="store_true",
        help="셀 값뿐 아니라 number format 차이도 변경으로 간주",
    )
    parser.add_argument(
        "--max-cell-details",
        type=int,
        default=2000,
        help="시트별 상세 변경 셀 최대 출력 수(기본: 2000)",
    )
    parser.add_argument(
        "--workers",
        type=int,
        default=default_worker_count(),
        help="폴더 비교 시 병렬 워커 수(기본: 자동, 최대 8)",
    )
    parser.add_argument(
        "--highlight-output",
        help="두번째 파일 기준으로 첫번째 파일과 다른 셀을 색상 표시한 결과 엑셀 파일 경로",
    )
    parser.add_argument(
        "--no-unprotect-second",
        action="store_true",
        help="두번째 파일 보호/숨김 해제 단계를 건너뜀(기본은 보호/숨김 해제 수행)",
    )
    return parser.parse_args(list(argv))


def main(argv: Iterable[str] | None = None) -> int:
    args = parse_args(argv if argv is not None else sys.argv[1:])
    left = Path(args.left).expanduser()
    right = Path(args.right).expanduser()

    if not left.exists():
        print(f"[ERROR] 경로를 찾을 수 없습니다: {left}", file=sys.stderr)
        return 2
    if not right.exists():
        print(f"[ERROR] 경로를 찾을 수 없습니다: {right}", file=sys.stderr)
        return 2

    if args.max_cell_details <= 0:
        print("[ERROR] --max-cell-details는 1 이상이어야 합니다.", file=sys.stderr)
        return 2
    if args.workers <= 0:
        print("[ERROR] --workers는 1 이상이어야 합니다.", file=sys.stderr)
        return 2

    highlight_output: Path | None = None
    if args.highlight_output:
        if not left.is_file() or not right.is_file():
            print("[ERROR] --highlight-output은 파일 vs 파일 비교에서만 사용할 수 있습니다.", file=sys.stderr)
            return 2
        if not is_supported_excel(left) or not is_supported_excel(right):
            print("[ERROR] --highlight-output 사용 시 양쪽 모두 xlsx/xlsm 파일이어야 합니다.", file=sys.stderr)
            return 2

        highlight_output = Path(args.highlight_output).expanduser()
        if not highlight_output.suffix:
            highlight_output = highlight_output.with_suffix(right.suffix)
        if highlight_output.suffix.lower() not in SUPPORTED_EXTENSIONS:
            print("[ERROR] 하이라이트 출력 파일 확장자는 xlsx 또는 xlsm 이어야 합니다.", file=sys.stderr)
            return 2

    try:
        report = compare_targets(
            left_target=left,
            right_target=right,
            include_format=args.include_format,
            max_cell_details=args.max_cell_details,
            workers=args.workers,
        )
    except ValueError as exc:
        print(f"[ERROR] {exc}", file=sys.stderr)
        return 2

    print_summary(report)

    if highlight_output:
        try:
            highlight_summary = generate_highlight_workbook(
                left_file=left,
                right_file=right,
                output_file=highlight_output,
                include_format=args.include_format,
                unprotect_second=not args.no_unprotect_second,
            )
        except (RuntimeError, OSError, BadZipFile) as exc:
            print(f"[ERROR] 하이라이트 파일 생성 실패: {exc}", file=sys.stderr)
            return 2

        print("=== Highlight Workbook ===")
        print(f"Output file  : {highlight_output.resolve()}")
        print(f"Modified cells: {highlight_summary.modified_cells}")
        print(f"Added cells   : {highlight_summary.added_cells}")
        print(f"Removed cells : {highlight_summary.removed_cells}")
        print(f"Changed sheets: {len(highlight_summary.changed_sheets)}")
        print(f"Added sheets  : {len(highlight_summary.added_sheets)}")
        print(f"Removed sheets: {len(highlight_summary.removed_sheets)}")
        if highlight_summary.changed_sheets:
            preview = ", ".join(highlight_summary.changed_sheets[:20])
            suffix = " ..." if len(highlight_summary.changed_sheets) > 20 else ""
            print(f"Changed sheet names: {preview}{suffix}")

    if args.output:
        output_path = Path(args.output).expanduser()
        output_format = args.format
        if output_path.suffix.lower() == ".json":
            output_format = "json"
        elif output_path.suffix.lower() == ".md":
            output_format = "md"

        output_path.parent.mkdir(parents=True, exist_ok=True)
        if output_format == "json":
            output_path.write_text(
                json.dumps(report_to_dict(report), ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
        else:
            output_path.write_text(render_markdown(report), encoding="utf-8")
        print(f"Report saved : {output_path.resolve()}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
